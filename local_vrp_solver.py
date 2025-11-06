import openpyxl             # permet de lire les fichiers excel ==> utilisé charger_donnees_excel pour ouvrir le fichier donnees_vrp.xlsx
import numpy as np          # pour les calculs mathématiques, surtout avec des matrices. C'est beaucoup plus rapide que les listes Python classiques pour ce genre de tâche.==> utilisé dans InstanceVRP pour calculer la distance euclidienne (le théorème de Pythagore)
import math                 # Fournit des fonctions mathématiques de base (sin, cos, exp, etc.). ==> utilisé dans resoudre_par_recuit_simule C'est le cœur du Recuit Simulé ! On l'utilise pour le calcul de "probabilité d'accepter une mauvaise solution".
import random               # C'est le module pour tout ce qui est aléatoire. Choisir au hasard, mélanger une liste, etc. ==> utilisé dans creer_solution_initiale (pour mélanger les clients. ) & creer_voisin_aleatoire (pour choisir entre "Relocate" et "Swap" et aussi pour choisir des clients au hasard.) & resoudre_par_recuit_simule (pour décider si on accepte ou non une moins bonne solution.)
import copy                 # Permet de faire de vraies copies (des doublons) des listes ou objets. C'est un point très important. ==> utilisé dans creer_voisin_aleatoire & resoudre_par_recuit_simule
import time                 # Permet de mesurer le temps. ==> utilisé dans resoudre_par_recuit_simule 
from pathlib import Path    # outil moderne pour gérer les chemins de fichiers (les "path") d'une manière qui marche sur tous les ordinateurs (Windows, Mac...).
import vrplib               # <-- NOUVEAU: Pour la validation scientifique
from urllib.error import URLError  # <-- NOUVEAU: Pour gérer les erreurs de téléchargement

# --- 1. Gestionnaire de Temps (Ta Contrainte) ---

class GestionnaireTempsTrajet:
    """
    Calcule le temps de trajet entre deux points en fonction
    de l'heure de départ et des multiplicateurs de trafic.
    """
    def __init__(self, matrice_distances_base, tranches_horaires_minutes, multiplicateurs):
        self.temps_trajet_base = matrice_distances_base
        self.tranches_horaires = tranches_horaires_minutes
        self.multiplicateurs = multiplicateurs

    def calculer_temps_trajet_et_arrivee(self, id_noeud_depart, id_noeud_arrivee, temps_depart_minutes):
        """
        Calcule le temps de trajet réel et l'heure d'arrivée.
        """
        # Trouver le bon multiplicateur de trafic
        multiplicateur_actuel = 1.0
        for i in range(len(self.tranches_horaires)):
            if temps_depart_minutes >= self.tranches_horaires[i]:
                multiplicateur_actuel = self.multiplicateurs[i]
            else:
                break
                
        temps_base = self.temps_trajet_base[id_noeud_depart, id_noeud_arrivee]
        temps_trajet_reel = temps_base * multiplicateur_actuel
        
        temps_arrivee_minutes = temps_depart_minutes + temps_trajet_reel
        
        return temps_trajet_reel, temps_arrivee_minutes

# --- 2. Classes de Données (Instance et Solution) ---

class InstanceVRP:
    """
    Stocke toutes les données du problème lues depuis l'Excel.
    (Capacité, demandes, coordonnées, etc.)
    """
    # Le paramètre 'matrice_explicite' a été supprimé
    def __init__(self, config, coords, demandes, noms):
        self.capacite_vehicule = config['CapaciteVehicule']
        self.nb_vehicules = config['NombreVehicules']
        self.id_depot = config['DepotID']
        
        self.coords_noeuds = coords
        self.demandes_clients = demandes
        self.noms_noeuds = noms
        
        self.ids_noeuds = sorted(list(self.coords_noeuds.keys()))
        self.nb_noeuds = len(self.ids_noeuds)
        
        self.ids_clients = [idx for idx in self.ids_noeuds if idx != self.id_depot]
        
        # Convertisseur entre l'ID (ex: 80) et l'index (ex: 3)
        self.id_vers_idx = {id_noeud: i for i, id_noeud in enumerate(self.ids_noeuds)}
        
        # On calcule toujours la matrice (Euclidienne) depuis l'Excel
        self.matrice_distances = np.zeros((self.nb_noeuds, self.nb_noeuds))
        for id1 in self.ids_noeuds:
            for id2 in self.ids_noeuds:
                if id1 == id2:
                    continue
                
                idx1 = self.id_vers_idx[id1]
                idx2 = self.id_vers_idx[id2]
                
                p1 = self.coords_noeuds[id1]
                p2 = self.coords_noeuds[id2]
                dist = np.linalg.norm(np.array(p1) - np.array(p2))
                
                self.matrice_distances[idx1, idx2] = dist

class Solution:
    """
    Représente une solution : une liste de tournées et son coût total.
    """
    def __init__(self, tournees, cout=float('inf')):
        self.tournees = tournees
        self.cout = cout 
        self.temps_calcul_interne = 0

    def calculer_cout(self, instance, gestionnaire_temps):
        """
        Calcule le coût (date de retour max) de la solution en tenant compte
        du trafic dynamique, de la capacité, et de la complétude.
        """
        temps_retour_max = 0
        solution_valide = True
        clients_servis = set()
        
        for tournee in self.tournees:
            temps_actuel = 0.0 # Départ du dépôt à t=0
            charge_actuelle = 0
            id_noeud_actuel = instance.id_depot
            
            tournee_valide = True # Vérifier chaque tournée
            
            for id_noeud_suivant in tournee[1:]: 
                
                # Ajouter le client au set (s'il n'est pas le dépôt)
                if id_noeud_suivant != instance.id_depot:
                    clients_servis.add(id_noeud_suivant)
                
                # 1. Vérifier la capacité
                charge_actuelle += instance.demandes_clients.get(id_noeud_suivant, 0)
                if charge_actuelle > instance.capacite_vehicule:
                    solution_valide = False # La solution GLOBALE est invalide
                    tournee_valide = False # Cette TOURNEE est invalide
                    break 
                
                # 2. Calculer le temps (Dynamique)
                idx_noeud_actuel = instance.id_vers_idx[id_noeud_actuel]
                idx_noeud_suivant = instance.id_vers_idx[id_noeud_suivant]
                
                temps_trajet, temps_arrivee = gestionnaire_temps.calculer_temps_trajet_et_arrivee(
                    idx_noeud_actuel, idx_noeud_suivant, temps_actuel
                )
                
                temps_actuel = temps_arrivee
                id_noeud_actuel = id_noeud_suivant
            
            # Mettre à jour le temps de retour (uniquement pour les tournées valides et non-vides)
            if tournee_valide and len(tournee) > 2 and temps_actuel > temps_retour_max:
                temps_retour_max = temps_actuel
        
        # 3. Vérifier que TOUS les clients sont servis
        # On fait cette vérification APRES avoir parcouru TOUTES les tournées
        if len(clients_servis) != len(instance.ids_clients):
            solution_valide = False
        
        if not solution_valide:
            self.cout = float('inf') 
            return self.cout
        
        self.cout = temps_retour_max
        return self.cout

# --- 3. Opérateurs de Voisinage ---

def creer_voisin_aleatoire(solution, instance):
    """
    Crée une nouvelle solution "voisine" en modifiant légèrement
    la solution actuelle (via Relocate ou Swap).
    """
    nouvelles_tournees = copy.deepcopy(solution.tournees)
    
    op = random.choice(['relocate', 'swap'])
    
    try:
        # Opérateur Relocate
        if op == 'relocate' and len(nouvelles_tournees) > 0:
            # Choisir une route non vide
            idx_tournee1 = random.choice([i for i, r in enumerate(nouvelles_tournees) if len(r) > 2])
            tournee1 = nouvelles_tournees[idx_tournee1]
            
            # Choisir un client à déplacer (ni le 1er ni le dernier)
            idx_client_a_bouger = random.randint(1, len(tournee1) - 2)
            client_a_bouger = tournee1.pop(idx_client_a_bouger)
            
            # Choisir une route de destination (peut être la même)
            idx_tournee2 = random.randint(0, len(nouvelles_tournees) - 1)
            tournee2 = nouvelles_tournees[idx_tournee2]
            
            # S'assurer que la tournée 2 n'est pas vide (juste [depot, depot])
            pos_insertion = 1
            if len(tournee2) > 2:
                pos_insertion = random.randint(1, len(tournee2) - 1)
            tournee2.insert(pos_insertion, client_a_bouger)

        # Opérateur Swap
        elif op == 'swap' and len(nouvelles_tournees) > 0:
            # Choisir 2 tournées (peuvent être les mêmes)
            idx_tournee1 = random.choice([i for i, r in enumerate(nouvelles_tournees) if len(r) > 2])
            idx_tournee2 = random.choice([i for i, r in enumerate(nouvelles_tournees) if len(r) > 2])
            tournee1 = nouvelles_tournees[idx_tournee1]
            tournee2 = nouvelles_tournees[idx_tournee2]

            # Choisir 2 clients
            idx_client1 = random.randint(1, len(tournee1) - 2)
            idx_client2 = random.randint(1, len(tournee2) - 2)

            # Échanger
            client1 = tournee1[idx_client1]
            client2 = tournee2[idx_client2]
            tournee1[idx_client1] = client2
            tournee2[idx_client2] = client1
    
    except (ValueError, IndexError):
        # Si une tournée est vide ou trop petite, on ne fait rien
        pass
        
    return Solution(nouvelles_tournees)

# --- 4. Création de la Solution Initiale ---

def creer_solution_initiale(instance):
    """
    Crée une solution de départ simple (First Fit).
    Essaie de créer une solution valide, mais si ce n'est pas possible,
    crée une solution invalide (que le Recuit devra réparer).
    """
    clients = copy.deepcopy(instance.ids_clients)
    random.shuffle(clients) 
    
    tournees = [[] for _ in range(instance.nb_vehicules)]
    charges_tournees = [0.0] * instance.nb_vehicules # Utiliser des floats
    
    clients_non_assignes = []

    # Logique "First Fit"
    for id_client in clients:
        if id_client not in instance.demandes_clients:
            print(f"Avertissement: Client ID {id_client} non trouvé dans les demandes.")
            continue
            
        demande_client = instance.demandes_clients[id_client]
        
        assigne = False
        for i in range(instance.nb_vehicules):
            if charges_tournees[i] + demande_client <= instance.capacite_vehicule:
                tournees[i].append(id_client)
                charges_tournees[i] += demande_client
                assigne = True
                break # Client assigné, passer au suivant
                
        if not assigne:
            # N'rentre dans aucun camion qui a de la place
            # On le "force" dans le camion le moins chargé
            idx_meilleure_tournee = np.argmin(charges_tournees)
            tournees[idx_meilleure_tournee].append(id_client)
            charges_tournees[idx_meilleure_tournee] += demande_client
            clients_non_assignes.append(id_client)
            
    if clients_non_assignes:
        print(f"\nAvertissement: {len(clients_non_assignes)} clients n'ont pas pu être assignés valablement et ont été forcés (capacité ?). L'algo va démarrer à 'inf'.")
            
    # Ajouter le dépôt au début et à la fin de chaque tournée
    tournees_finales = []
    for tournee in tournees:
        tournees_finales.append([instance.id_depot] + tournee + [instance.id_depot])
        
    # S'assurer qu'on a le bon nombre de véhicules
    while len(tournees_finales) < instance.nb_vehicules:
        tournees_finales.append([instance.id_depot, instance.id_depot])
        
    return Solution(tournees_finales)

# --- 5. Algorithme de Recuit Simulé ---

def resoudre_par_recuit_simule(instance, gestionnaire_temps, temp_initiale, temp_finale, taux_refroidissement, iter_par_temp, verbose=True):
    """
    C'est le "cerveau" : il explore les solutions en utilisant
    l'algorithme du Recuit Simulé pour trouver la meilleure.
    """
    temps_debut = time.time()
    
    solution_actuelle = creer_solution_initiale(instance)
    cout_actuel = solution_actuelle.calculer_cout(instance, gestionnaire_temps)
    
    meilleure_solution = copy.deepcopy(solution_actuelle)
    meilleur_cout = cout_actuel
    
    temp = temp_initiale
    
    if verbose:
        print(f"Coût initial: {cout_actuel:.2f}")
    
    while temp > temp_finale:
        for _ in range(iter_par_temp):
            
            nouvelle_solution = creer_voisin_aleatoire(solution_actuelle, instance)
            nouveau_cout = nouvelle_solution.calculer_cout(instance, gestionnaire_temps)
            
            # Logique d'acceptation de Metropolis (gère 'inf')
            
            if nouveau_cout < cout_actuel:
                # Cas 1: C'est une amélioration (30000 < 31000, ou 30000 < inf)
                solution_actuelle = nouvelle_solution
                cout_actuel = nouveau_cout
            
            elif nouveau_cout > cout_actuel and cout_actuel != float('inf'):
                # Cas 2: C'est une dégradation (31000 > 30000) ET on part d'une solution valide
                delta_cout = nouveau_cout - cout_actuel
                if random.random() < math.exp(-delta_cout / temp):
                    solution_actuelle = nouvelle_solution
                    cout_actuel = nouveau_cout
            
            # Mettre à jour la meilleure solution jamais vue
            if nouveau_cout < meilleur_cout: 
                meilleure_solution = copy.deepcopy(nouvelle_solution)
                meilleur_cout = nouveau_cout
                    
        temp *= taux_refroidissement
        
    temps_fin = time.time()
    temps_total = temps_fin - temps_debut
    
    if verbose:
        print(f"\nRecuit terminé.")
        print(f"Meilleur coût (temps retour max): {meilleur_cout:.2f} minutes")
        print(f"Temps de calcul: {temps_total:.2f}s")
    
    meilleure_solution.temps_calcul_interne = temps_total
    
    return meilleure_solution, meilleur_cout

# --- 6. Chargement des Données Excel ---

def charger_donnees_excel(chemin_fichier):
    """
    Ouvre le fichier Excel et lit les 3 feuilles :
    'Config', 'Coordonnees', 'Demandes'.
    """
    try:
        classeur = openpyxl.load_workbook(chemin_fichier, data_only=True)
    except FileNotFoundError:
        print(f"ERREUR: Fichier non trouvé '{chemin_fichier}'.")
        print("Veuillez créer le fichier et le placer à côté du script.")
        exit()

    print(f"Lecture du fichier '{chemin_fichier}'...")
    
    try:
        # Feuille 'Config'
        config = {}
        feuille_config = classeur['Config']
        for ligne in feuille_config.iter_rows(min_row=2, values_only=True):
            if ligne[0] is not None and ligne[1] is not None:
                config[ligne[0]] = int(ligne[1])

        # Feuille 'Coordonnees'
        coords = {}
        noms = {}
        feuille_coords = classeur['Coordonnees']
        for ligne in feuille_coords.iter_rows(min_row=2, values_only=True):
            if ligne[0] is not None and ligne[1] is not None and ligne[2] is not None:
                id_noeud = int(ligne[0])
                coords[id_noeud] = (float(ligne[1]), float(ligne[2]))
                # Gérer le cas où la colonne Nom (D) est vide
                noms[id_noeud] = str(ligne[3]) if ligne[3] is not None else f"ID {id_noeud}"

        # Feuille 'Demandes'
        demandes = {}
        feuille_demandes = classeur['Demandes']
        for ligne in feuille_demandes.iter_rows(min_row=2, values_only=True):
            if ligne[0] is not None and ligne[1] is not None:
                id_noeud = int(ligne[0])
                demandes[id_noeud] = int(ligne[1])

        print("Données chargées avec succès.")
        return config, coords, demandes, noms

    except (ValueError, TypeError, IndexError) as e:
        print(f"\nERREUR: Problème de format de données dans le fichier Excel.")
        print(f"Vérifiez que les colonnes ID, X, Y, Demande, etc. sont correctes.")
        print(f"Détail de l'erreur: {e}")
        exit()

# --- 7. Exécution Principale (Mode Excel Uniquement) ---

if __name__ == "__main__":
    
    print("--- Démarrage du Solveur VRP (Mode Excel & Trafic Dynamique) ---")
    
    # === A. Chargement des données locales ===
    DOSSIER_SCRIPT = Path(__file__).parent
    # On utilise ton fichier Excel avec 200 clients
    FICHIER_EXCEL = DOSSIER_SCRIPT / "donnees_vrp_200.xlsx" 
    config, coords, demandes, noms = charger_donnees_excel(FICHIER_EXCEL)
    
    instance_vrp = InstanceVRP(config, coords, demandes, noms)
    
    print("\n--- Configuration de l'instance ---")
    print(f"  - Dépôt (ID): {instance_vrp.id_depot} ({instance_vrp.noms_noeuds[instance_vrp.id_depot]})")
    print(f"  - Nombre de clients: {len(instance_vrp.ids_clients)}")
    print(f"  - Nb. Véhicules: {instance_vrp.nb_vehicules}")
    print(f"  - Capacité / Véhicule: {instance_vrp.capacite_vehicule}")
    
    # === B. Configuration de la contrainte (Trafic Dynamique) ===
    # C'est ta contrainte principale !
    tranches = [0, 8 * 60, 10 * 60, 16 * 60, 18 * 60] # 0h, 8h, 10h, 16h, 18h
    multiplicateurs = [1.0, 1.8, 1.2, 2.0, 1.0] # Normal, Pointe Matin, Creux, Pointe Soir, Normal
    
    gestionnaire_temps = GestionnaireTempsTrajet(
        instance_vrp.matrice_distances, tranches, multiplicateurs
    )
    print("\n--- Configuration du Trafic ---")
    print(f"Tranches horaires (minutes): {tranches}")
    print(f"Multiplicateurs de trafic: {multiplicateurs}")
    
    # === C. Configuration du Recuit Simulé ===
    TEMPERATURE_INITIALE = 1000.0 # Plus haut pour un plus gros problème
    TEMPERATURE_FINALE = 0.1
    TAUX_REFROIDISSEMENT = 0.99 # Refroidissement lent
    ITERATIONS_PAR_PALIER = 200 # Plus d'itérations
    
    NB_LANCES = 20 # Le "20 runs" du projet
    
    # === D. Lancement de l'étude statistique ===
    
    print(f"\n--- Lancement de l'Étude Statistique ({NB_LANCES} lancements) ---")
    
    tous_les_couts = []
    meilleure_solution_globale = None
    meilleur_cout_global = float('inf')
    temps_total_etude = 0
    
    for i in range(NB_LANCES):
        print(f"  Lancement {i+1}/{NB_LANCES}...", end="", flush=True) 
        
        solution_finale, cout_final = resoudre_par_recuit_simule(
            instance_vrp,
            gestionnaire_temps,
            TEMPERATURE_INITIALE,
            TEMPERATURE_FINALE,
            TAUX_REFROIDISSEMENT,
            ITERATIONS_PAR_PALIER,
            verbose=False # On désactive les prints dans la boucle
        )
        
        tous_les_couts.append(cout_final)
        temps_total_etude += solution_finale.temps_calcul_interne
        
        if cout_final < meilleur_cout_global:
            meilleur_cout_global = cout_final
            meilleure_solution_globale = solution_finale
            
        print(f" Coût trouvé: {cout_final:.2f}")

    print(f"\nTemps total de l'étude: {temps_total_etude:.2f}s (Moyenne: {(temps_total_etude/NB_LANCES):.2f}s / run)")
    
    # === E. Affichage des résultats statistiques ===
    print("\n--- Résultats de l'Étude Statistique ---")
    
    resultats_couts = np.array(tous_les_couts)
    # Gérer le cas où tous les coûts sont 'inf'
    if np.all(resultats_couts == float('inf')):
        print("  Meilleur Coût: inf")
        print("  Pire Coût:       inf")
        print("  Coût Moyen:     inf")
        print("  Écart-type:     nan (mesure la stabilité de l'algo)")
    else:
        # Filtrer les 'inf' pour les statistiques
        couts_valides = resultats_couts[resultats_couts != float('inf')]
        if len(couts_valides) > 0:
            print(f"  Meilleur Coût: {np.min(couts_valides):.2f}")
            print(f"  Pire Coût:       {np.max(couts_valides):.2f}")
            print(f"  Coût Moyen:     {np.mean(couts_valides):.2f}")
            print(f"  Écart-type:     {np.std(couts_valides):.2f} (mesure la stabilité de l'algo)")
        else:
            print("  Aucune solution valide trouvée (0 coût valide / 20 lancements).")

    # === G. Affichage de la meilleure solution ===
    print("\n--- Meilleure Solution Trouvée (sur les 20 lancements) ---")
    print(f"Le meilleur coût global trouvé est : {meilleur_cout_global:.2f}")
    print("Tournées (listes de noms):")
    
    if meilleure_solution_globale is not None and meilleur_cout_global != float('inf'):
        for i, tournee in enumerate(meilleure_solution_globale.tournees):
            noms_tournee = [instance_vrp.noms_noeuds[id_noeud] for id_noeud in tournee]
            if len(noms_tournee) > 10:
                print(f"  Véhicule {i+1} ({len(noms_tournee)-2} clients): {noms_tournee[0]} -> {noms_tournee[1]} -> ... -> {noms_tournee[-2]} -> {noms_tournee[-1]}")
            elif len(noms_tournee) > 2: # N'affiche pas les tournées vides
                print(f"  Véhicule {i+1}: {' -> '.join(noms_tournee)}")
    else:
        print("Aucune solution valide n'a été trouvée.")

