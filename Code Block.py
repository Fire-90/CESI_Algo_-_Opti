# --- Imports Nécessaires --- 
import openpyxl
import pandas as pd
import numpy as np
import math
import random
import copy
import time
from pathlib import Path

# --- Imports pour les graphiques ---
import matplotlib.pyplot as plt
import seaborn as sns

def generer_et_sauvegarder_instance(nb_clients, capacite_vehicule, nb_vehicules, dossier_sortie, nom_fichier_base):
    nom_fichier = dossier_sortie / f"{nom_fichier_base}_{nb_clients}_clients.xlsx"
    ids = list(range(1, nb_clients + 2)); depot_id = 1
    coords_x = np.random.randint(0, 101, size=nb_clients + 1); coords_y = np.random.randint(0, 101, size=nb_clients + 1)
    noms = [f"Client-{i}" for i in ids]; noms[ids.index(depot_id)] = "Paris"
    df_coords = pd.DataFrame({'ID': ids, 'X': coords_x, 'Y': coords_y, 'Nom': noms, 'info_supp': ['(dépôt)'] + [f'client {i-1}' for i in range(2, nb_clients + 2)]})
    demandes = np.random.randint(5, 21, size=nb_clients + 1); demandes[ids.index(depot_id)] = 0
    df_demandes = pd.DataFrame({'ID': ids, 'Demande': demandes, 'info_supp': ['le dépôt n\'a aucune commande'] + [f'le client {i-1} commande {demandes[i-1]} unités' for i in range(2, nb_clients + 2)]})
    df_config = pd.DataFrame({'Propriete': ['CapaciteVehicule', 'NombreVehicules', 'DepotID'], 'Valeur': [capacite_vehicule, nb_vehicules, depot_id], 'info_supp': ['Capacite Max du véhicule', 'Nb de véhicules disponibles', 'ID du point de départ']})
    with pd.ExcelWriter(nom_fichier, engine='openpyxl') as writer:
        df_coords.to_excel(writer, sheet_name='Coordonnees', index=False); df_demandes.to_excel(writer, sheet_name='Demandes', index=False); df_config.to_excel(writer, sheet_name='Config', index=False)
    return nom_fichier

class GestionnaireTempsTrajet:
    def __init__(self, matrice_distances_base, tranches_horaires_minutes, multiplicateurs):
        self.temps_trajet_base = matrice_distances_base; self.tranches_horaires = tranches_horaires_minutes; self.multiplicateurs = multiplicateurs
    def calculer_temps_trajet_et_arrivee(self, id_noeud_depart, id_noeud_arrivee, temps_depart_minutes):
        multiplicateur_actuel = 1.0
        for i in range(len(self.tranches_horaires)):
            if temps_depart_minutes >= self.tranches_horaires[i]: multiplicateur_actuel = self.multiplicateurs[i]
            else: break
        temps_base = self.temps_trajet_base[id_noeud_depart, id_noeud_arrivee]; temps_trajet_reel = temps_base * multiplicateur_actuel
        temps_arrivee_minutes = temps_depart_minutes + temps_trajet_reel; return temps_trajet_reel, temps_arrivee_minutes
class InstanceVRP:
    def __init__(self, config, coords, demandes, noms):
        self.capacite_vehicule = config['CapaciteVehicule']; self.nb_vehicules = config['NombreVehicules']; self.id_depot = config['DepotID']
        self.coords_noeuds = coords; self.demandes_clients = demandes; self.noms_noeuds = noms
        self.ids_noeuds = sorted(list(self.coords_noeuds.keys())); self.nb_noeuds = len(self.ids_noeuds)
        self.ids_clients = [idx for idx in self.ids_noeuds if idx != self.id_depot]
        self.id_vers_idx = {id_noeud: i for i, id_noeud in enumerate(self.ids_noeuds)}; self.matrice_distances = self._calculer_matrice_distances_euclidienne()
    def _calculer_matrice_distances_euclidienne(self):
        matrice = np.zeros((self.nb_noeuds, self.nb_noeuds))
        for id1 in self.ids_noeuds:
            for id2 in self.ids_noeuds:
                idx1, idx2 = self.id_vers_idx[id1], self.id_vers_idx[id2]
                p1, p2 = self.coords_noeuds[id1], self.coords_noeuds[id2]
                matrice[idx1, idx2] = np.linalg.norm(np.array(p1) - np.array(p2))
        return matrice
class Solution:
    def __init__(self, tournees, cout=float('inf')):
        self.tournees = tournees; self.cout = cout; self.temps_calcul = 0
    def calculer_cout_total(self, instance, gestionnaire_temps):
        temps_retour_max = 0; solution_valide = True; clients_servis = set()
        for tournee in self.tournees:
            temps_actuel = 0.0; charge_actuelle = 0; id_noeud_actuel = instance.id_depot
            for id_noeud_suivant in tournee[1:]:
                if id_noeud_suivant != instance.id_depot: clients_servis.add(id_noeud_suivant)
                charge_actuelle += instance.demandes_clients.get(id_noeud_suivant, 0)
                if charge_actuelle > instance.capacite_vehicule: solution_valide = False; break
                idx_actuel, idx_suivant = instance.id_vers_idx[id_noeud_actuel], instance.id_vers_idx[id_noeud_suivant]
                _, temps_arrivee = gestionnaire_temps.calculer_temps_trajet_et_arrivee(idx_actuel, idx_suivant, temps_actuel)
                temps_actuel, id_noeud_actuel = temps_arrivee, id_noeud_suivant
            if not solution_valide: break
            if len(tournee) > 2 and temps_actuel > temps_retour_max: temps_retour_max = temps_actuel
        if len(clients_servis) != len(instance.ids_clients): solution_valide = False
        self.cout = temps_retour_max if solution_valide else float('inf')
        return self.cout
    
class DataLoader:
    @staticmethod
    def charger_donnees_excel(chemin_fichier):
        try: classeur = openpyxl.load_workbook(chemin_fichier, data_only=True)
        except FileNotFoundError: print(f"ERREUR: Fichier non trouvé '{chemin_fichier}'."); exit()
        try:
            config = {l[0]: int(l[1]) for l in classeur['Config'].iter_rows(min_row=2, values_only=True) if l[0]}
            coords = {int(l[0]): (float(l[1]), float(l[2])) for l in classeur['Coordonnees'].iter_rows(min_row=2, values_only=True) if l[0]}
            noms = {int(l[0]): str(l[3]) if l[3] else f"ID {l[0]}" for l in classeur['Coordonnees'].iter_rows(min_row=2, values_only=True) if l[0]}
            demandes = {int(l[0]): int(l[1]) for l in classeur['Demandes'].iter_rows(min_row=2, values_only=True) if l[0]}
            return config, coords, demandes, noms
        except (ValueError, TypeError, IndexError) as e: print(f"\nERREUR de format dans le fichier Excel: {e}"); exit()

class SimulatedAnnealingSolverOptimized:
    def __init__(self, instance, gest_tps, temp_i, temp_f, taux_r, iter_p_t):
        self.instance, self.gest_tps, self.temp_i, self.temp_f, self.taux_r, self.iter_p_t = instance, gest_tps, temp_i, temp_f, taux_r, iter_p_t
    def _creer_solution_initiale(self):
        clients = list(self.instance.ids_clients); random.shuffle(clients)
        tournees = [[] for _ in range(self.instance.nb_vehicules)]; charges = [0.0] * self.instance.nb_vehicules
        for id_client in clients:
            demande = self.instance.demandes_clients[id_client]; assigne = False
            for i in range(self.instance.nb_vehicules):
                if charges[i] + demande <= self.instance.capacite_vehicule:
                    tournees[i].append(id_client); charges[i] += demande; assigne = True; break
            if not assigne: return None
        for tour in tournees: random.shuffle(tour)
        return Solution([[self.instance.id_depot] + t + [self.instance.id_depot] for t in tournees])
    def _calculer_cout_tournee(self, tournee):
        tps, charge, noeud = 0.0, 0, self.instance.id_depot
        for noeud_suiv in tournee[1:]:
            charge += self.instance.demandes_clients.get(noeud_suiv, 0)
            if charge > self.instance.capacite_vehicule: return float('inf')
            idx_a, idx_s = self.instance.id_vers_idx[noeud], self.instance.id_vers_idx[noeud_suiv]
            _, tps_arr = self.gest_tps.calculer_temps_trajet_et_arrivee(idx_a, idx_s, tps)
            tps, noeud = tps_arr, noeud_suiv
        return tps
    def solve(self, verbose=False):
        t_start = time.time()
        sol_act = None
        for _ in range(50):
            sol_act = self._creer_solution_initiale()
            if sol_act: break
        if not sol_act:
            sol_invalide = Solution([], float('inf')); sol_invalide.temps_calcul = time.time() - t_start
            return sol_invalide
        couts_tournees_act = [self._calculer_cout_tournee(t) for t in sol_act.tournees]
        cout_act = max(c for c in couts_tournees_act) if any(c != float('inf') for c in couts_tournees_act) else float('inf')
        meilleure_sol, meilleur_cout = copy.deepcopy(sol_act), cout_act
        temp = self.temp_i
        while temp > self.temp_f and cout_act != float('inf'):
            for _ in range(self.iter_p_t):
                tournees = sol_act.tournees
                mod_idx = [i for i, r in enumerate(tournees) if len(r) > 2]
                if not mod_idx: continue
                annul_info, a_recalculer = {}, set()
                op = random.choice(['relocate', 'swap', 'swap_intra'])
                try:
                    if op == 'relocate':
                        idx_t1 = random.choice(mod_idx)
                        idx_c1 = random.randint(1, len(tournees[idx_t1]) - 2)
                        client = tournees[idx_t1].pop(idx_c1)
                        idx_t2 = random.randrange(len(tournees))
                        pos_ins = random.randint(1, len(tournees[idx_t2]) - 1) if len(tournees[idx_t2]) > 2 else 1
                        tournees[idx_t2].insert(pos_ins, client)
                        annul_info = {'op': 'relocate', 'client': client, 'from_t': idx_t1, 'from_c': idx_c1, 'to_t': idx_t2, 'to_c': pos_ins}
                        a_recalculer = {idx_t1, idx_t2}
                    elif op == 'swap' and len(mod_idx) >= 2:
                        idx_t1, idx_t2 = random.sample(mod_idx, 2)
                        idx_c1, idx_c2 = random.randint(1, len(tournees[idx_t1]) - 2), random.randint(1, len(tournees[idx_t2]) - 2)
                        c1, c2 = tournees[idx_t1][idx_c1], tournees[idx_t2][idx_c2]
                        tournees[idx_t1][idx_c1], tournees[idx_t2][idx_c2] = c2, c1
                        annul_info = {'op': 'swap', 't1': idx_t1, 'c1_idx': idx_c1, 't2': idx_t2, 'c2_idx': idx_c2, 'c1': c1, 'c2': c2}
                        a_recalculer = {idx_t1, idx_t2}
                    else: # swap_intra
                        idx_t1 = random.choice(mod_idx)
                        if len(tournees[idx_t1]) > 3:
                           idx_c1, idx_c2 = random.sample(range(1, len(tournees[idx_t1]) - 1), 2)
                           c1, c2 = tournees[idx_t1][idx_c1], tournees[idx_t1][idx_c2]
                           tournees[idx_t1][idx_c1], tournees[idx_t1][idx_c2] = c2, c1
                           annul_info = {'op': 'swap_intra', 't1': idx_t1, 'c1_idx': idx_c1, 'c2_idx': idx_c2, 'c1': c1, 'c2': c2}
                           a_recalculer = {idx_t1}
                    if not a_recalculer: continue
                    couts_tournees_voisin = list(couts_tournees_act)
                    for i in a_recalculer: couts_tournees_voisin[i] = self._calculer_cout_tournee(tournees[i])
                    cout_voisin = max(c for c in couts_tournees_voisin) if any(c != float('inf') for c in couts_tournees_voisin) else float('inf')
                    if cout_voisin < cout_act or random.random() < math.exp(-(cout_voisin - cout_act) / temp):
                        cout_act, couts_tournees_act = cout_voisin, couts_tournees_voisin
                        if cout_act < meilleur_cout:
                            meilleur_cout = cout_act
                            meilleure_sol = Solution(copy.deepcopy(tournees), meilleur_cout)
                    else:
                        if annul_info.get('op') == 'relocate':
                            tournees[annul_info['to_t']].pop(annul_info['to_c']); tournees[annul_info['from_t']].insert(annul_info['from_c'], annul_info['client'])
                        elif annul_info.get('op') == 'swap':
                            tournees[annul_info['t1']][annul_info['c1_idx']], tournees[annul_info['t2']][annul_info['c2_idx']] = annul_info['c1'], annul_info['c2']
                        elif annul_info.get('op') == 'swap_intra':
                            tournees[annul_info['t1']][annul_info['c1_idx']], tournees[annul_info['t1']][annul_info['c2_idx']] = annul_info['c1'], annul_info['c2']
                except (ValueError, IndexError): pass
            temp *= self.taux_r
        meilleure_sol.temps_calcul = time.time() - t_start
        meilleure_sol.cout = meilleur_cout
        return meilleure_sol
    
DOSSIER_SCRIPT = Path.cwd(); DOSSIER_INSTANCES = DOSSIER_SCRIPT / "instances_generees"; DOSSIER_INSTANCES.mkdir(exist_ok=True)
TAILLES_INSTANCES = [500]
NB_LANCES_RS = 5
CAPACITE_VEHICULE = 150
TEMP_I, TEMP_F, TAUX_R = 1000.0, 0.1, 0.99

# Dictionnaire pour stocker tous les résultats
resultats_pour_graphiques = {}

for nb_clients in TAILLES_INSTANCES:
    dem_moy = 15; dem_tot_est = nb_clients * dem_moy; nb_veh_min = math.ceil(dem_tot_est / CAPACITE_VEHICULE)
    NB_VEHICULES = int(nb_veh_min * 1.4) + 2
    ITER_P_T = 1000 if nb_clients < 500 else 500

    fichier = generer_et_sauvegarder_instance(nb_clients, CAPACITE_VEHICULE, NB_VEHICULES, DOSSIER_INSTANCES, f"instance_{nb_clients}")
    config, coords, demandes, noms = DataLoader.charger_donnees_excel(fichier)
    instance = InstanceVRP(config, coords, demandes, noms)
    gest_tps = GestionnaireTempsTrajet(instance.matrice_distances, [0, 8*60, 10*60, 16*60, 18*60], [1.0, 1.8, 1.2, 2.0, 1.0])
    
    print(f"\n\n{'='*25} INSTANCE: {nb_clients} CLIENTS {'='*25}")
    print("\n--- Configuration de l'instance ---")
    print(f"  - Dépôt (ID): {instance.id_depot} ({instance.noms_noeuds[instance.id_depot]})")
    print(f"  - Nombre de clients: {len(instance.ids_clients)}")
    print(f"  - Nb. Véhicules: {instance.nb_vehicules}")
    print(f"  - Capacité / Véhicule: {instance.capacite_vehicule}")
    print("\n--- Configuration du Trafic ---")
    print(f"Tranches horaires (minutes): {[0, 480, 600, 960, 1080]}")
    print(f"Multiplicateurs de trafic: {[1.0, 1.8, 1.2, 2.0, 1.0]}")
    
    print(f"\n--- Lancement de l'Étude Statistique ({NB_LANCES_RS} lancements) ---")
    solveur = SimulatedAnnealingSolverOptimized(instance, gest_tps, TEMP_I, TEMP_F, TAUX_R, ITER_P_T)
    all_costs, best_cost, best_sol = [], float('inf'), None
    t_start_study = time.time()
    for i in range(NB_LANCES_RS):
        solution = solveur.solve(verbose=False)
        all_costs.append(solution.cout)
        if solution.cout < best_cost: best_cost, best_sol = solution.cout, solution
        print(f"  Lancement {i+1}/{NB_LANCES_RS}... Coût trouvé: {solution.cout:.2f}")
    total_time = time.time() - t_start_study
    print(f"\nTemps total de l'étude: {total_time:.2f}s (Moyenne: {(total_time/NB_LANCES_RS):.2f}s / run)")
    
    print("\n--- Résultats de l'Étude Statistique ---")
    valid_costs = [c for c in all_costs if c != float('inf')]
    if valid_costs:
        best_cost_rs = np.min(valid_costs)
        print(f"  Meilleur Coût: {best_cost_rs:.2f}")
        print(f"  Pire Coût:       {np.max(valid_costs):.2f}")
        print(f"  Coût Moyen:     {np.mean(valid_costs):.2f}")
        print(f"  Écart-type:     {np.std(valid_costs):.2f} (mesure la stabilité)")
    else: print("  Aucune solution valide n'a été trouvée.")
    
    print("\n--- Meilleure Solution Trouvée ---")
    if best_sol and best_cost != float('inf'):
        print(f"Le meilleur coût global trouvé est : {best_cost:.2f}")
        for i, tournee in enumerate(best_sol.tournees):
            if len(tournee) > 2:
                noms_tournee = [instance.noms_noeuds[id_noeud] for id_noeud in tournee]
                print(f"  Véhicule {i+1} ({len(noms_tournee)-2} clients): {' -> '.join(noms_tournee)}")
    else:
        print("Aucune solution valide à afficher.")
# Sauvegarde des résultats pour cette instance
resultats_pour_graphiques[nb_clients] = {'costs': all_costs, 'num_runs': NB_LANCES_RS}

for nb_clients, data in resultats_pour_graphiques.items():
    
    all_costs = data['costs']
    NB_LANCES = data['num_runs']

    # 1. Préparer les données en filtrant les solutions invalides
    couts_valides = [c for c in all_costs if c != float('inf')]

    # 2. Vérifier s'il y a des données à afficher
    if not couts_valides:
        print(f"\nInstance {nb_clients} clients: Aucune solution valide, impossible de générer le graphique.")
        continue # Passe à l'instance suivante
    
    # 3. Créer la figure et le graphique
    plt.style.use('seaborn-v0_8-whitegrid')
    fig, ax = plt.subplots(figsize=(8, 6))

    ax.boxplot(couts_valides, vert=True, patch_artist=True, widths=0.4)

    # 4. Calculer les métriques
    moyenne = np.mean(couts_valides)
    minimum = np.min(couts_valides)
    
    # MODIFICATION: Calculer le GAP pour l'affichage
    gap_graph = 0.0
    if minimum > 0:
        gap_graph = 100 * (moyenne - minimum) / minimum
    
    # 5. Améliorer l'affichage
    ax.set_title(f'Distribution des Coûts sur {NB_LANCES} Lancements\n(Instance à {nb_clients} clients)', fontsize=16)
    ax.set_ylabel('Coût de la Solution', fontsize=12)
    ax.set_xticklabels(['Recuit Simulé'])
    ax.yaxis.grid(True, linestyle='--', which='major', color='grey', alpha=0.7)
    
    # Tracer la moyenne
    ax.scatter(1, moyenne, marker='o', color='red', s=50, zorder=3, label=f'Moyenne: {moyenne:.2f}')
    ax.legend()
    
    # MODIFICATION: Ajouter le texte du GAP sur le graphique
    ax.text(0.95, 0.95, f'GAP (Moy/Min): {gap_graph:.2f}%',
            transform=ax.transAxes,
            fontsize=12,
            verticalalignment='top',
            horizontalalignment='right',
            bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
    
    # 6. Afficher le graphique final pour cette instance
    print(f"Affichage du graphique pour {nb_clients} clients...")
    plt.show()